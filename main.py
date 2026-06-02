"""
Money Forward クラウド勤怠のエクスポートファイル（CSV または Excel）を読み込み、
Google スプレッドシートへ店舗ごとに書き込む。

使用方法:
    python main.py <ファイルパス>

例:
    python main.py downloads/attendance_202405.csv
    python main.py sample_data.xlsx
"""

import sys
import re
import csv
from pathlib import Path

import openpyxl
import gspread
from google.oauth2.service_account import Credentials

# ---- 設定（環境に合わせて変更） ----
SPREADSHEET_ID = "YOUR_SPREADSHEET_ID"  # スプレッドシートのIDに置き換える
CREDENTIALS_FILE = Path(__file__).parent / "credentials" / "service_account.json"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Money Forward エクスポートの列名
COL_STORE = "所属"
COL_NAME = "氏名"
COL_OVERTIME = "時間外労働時間"


# ---- Excel 読み込み ----

def _parse_time_to_minutes(value) -> int:
    """HH:MM 形式または数値を分に変換する"""
    if value is None or str(value).strip() == "":
        return 0
    s = str(value).strip()
    m = re.match(r"^(\d+):(\d{2})$", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        return int(float(s) * 60)
    except ValueError:
        return 0


def minutes_to_hhmm(minutes: int) -> str:
    return f"{minutes // 60}:{minutes % 60:02d}"


def _parse_rows(headers: list[str], rows: list[list]) -> list[dict]:
    """ヘッダーと行データからレコード一覧を生成する"""
    def require_col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            raise ValueError(f"列 '{name}' が見つかりません。列一覧: {headers}")

    idx_store = require_col(COL_STORE)
    idx_name = require_col(COL_NAME)
    idx_overtime = require_col(COL_OVERTIME)

    records = []
    for row in rows:
        if not any(row):
            continue
        store = str(row[idx_store]).strip() if row[idx_store] else ""
        name = str(row[idx_name]).strip() if row[idx_name] else ""
        if not store or not name:
            continue
        minutes = _parse_time_to_minutes(row[idx_overtime])
        records.append({
            "store": store,
            "name": name,
            "overtime_minutes": minutes,
            "overtime_hhmm": minutes_to_hhmm(minutes),
        })
    return records


def read_file(filepath: str) -> list[dict]:
    """CSV または Excel を読み込み、店舗・氏名・残業時間のレコード一覧を返す"""
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        # CSV 読み込み（BOM付きUTF-8 / Shift-JIS 両対応）
        for encoding in ("utf-8-sig", "shift-jis", "utf-8"):
            try:
                with open(path, newline="", encoding=encoding) as f:
                    reader = csv.reader(f)
                    all_rows = list(reader)
                if all_rows:
                    headers = [h.strip() for h in all_rows[0]]
                    return _parse_rows(headers, all_rows[1:])
            except (UnicodeDecodeError, StopIteration):
                continue
        raise ValueError(f"CSVファイルのエンコーディングを判定できませんでした: {filepath}")

    else:
        # Excel 読み込み (.xlsx / .xls)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return _parse_rows(headers, rows)


# ---- Google Sheets 書き込み ----

def _get_or_create_sheet(ss: gspread.Spreadsheet, title: str) -> gspread.Worksheet:
    existing = {ws.title for ws in ss.worksheets()}
    if title in existing:
        return ss.worksheet(title)
    return ss.add_worksheet(title=title, rows=200, cols=10)


def write_to_sheets(records: list[dict]):
    """店舗別シート＋集計シートへ書き込む"""
    creds = Credentials.from_service_account_file(str(CREDENTIALS_FILE), scopes=SCOPES)
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(SPREADSHEET_ID)

    # 店舗ごとにグループ化
    by_store: dict[str, list[dict]] = {}
    for rec in records:
        by_store.setdefault(rec["store"], []).append(rec)

    # 店舗別シートへ書き込み
    for store, employees in sorted(by_store.items()):
        ws = _get_or_create_sheet(ss, store)
        rows = [["氏名", "残業時間", "残業時間（分）"]]
        for emp in sorted(employees, key=lambda x: x["name"]):
            rows.append([emp["name"], emp["overtime_hhmm"], emp["overtime_minutes"]])
        ws.clear()
        ws.update(rows, "A1")
        print(f"  ✓ {store}: {len(employees)} 名")

    # 集計シートへ書き込み
    summary_ws = _get_or_create_sheet(ss, "集計")
    summary_rows = [["店舗", "氏名", "残業時間", "残業時間（分）"]]
    for rec in sorted(records, key=lambda x: (x["store"], x["name"])):
        summary_rows.append([
            rec["store"], rec["name"],
            rec["overtime_hhmm"], rec["overtime_minutes"],
        ])
    summary_ws.clear()
    summary_ws.update(summary_rows, "A1")
    print(f"  ✓ 集計シート: {len(records)} 件")


# ---- エントリポイント ----

def main():
    if len(sys.argv) < 2:
        print("使用方法: python main.py <Excelファイルパス>")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not Path(excel_path).exists():
        print(f"エラー: ファイルが見つかりません: {excel_path}")
        sys.exit(1)

    print(f"読み込み中: {excel_path}")
    records = read_file(excel_path)
    print(f"  → {len(records)} 件のデータを読み込みました\n")

    print("スプレッドシートへ書き込み中...")
    write_to_sheets(records)
    print("\n完了！")


if __name__ == "__main__":
    main()
