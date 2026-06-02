"""
動作確認用のサンプルExcelを生成するスクリプト。
Money Forward クラウド勤怠のエクスポート形式を模倣。

使用方法:
    python create_sample.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


SAMPLE_DATA = [
    # (所属, 氏名, 時間外労働時間)
    ("渋谷店", "田中 一郎", "12:30"),
    ("渋谷店", "鈴木 花子", "8:00"),
    ("渋谷店", "佐藤 次郎", "20:15"),
    ("渋谷店", "高橋 美咲", "5:45"),
    ("新宿店", "伊藤 健太", "15:00"),
    ("新宿店", "渡辺 由美", "3:30"),
    ("新宿店", "山本 大輔", "22:00"),
    ("新宿店", "中村 さくら", "10:30"),
    ("池袋店", "小林 誠", "18:45"),
    ("池袋店", "加藤 奈々", "6:15"),
    ("池袋店", "吉田 翔太", "0:00"),
]

HEADER_ROW = [
    "所属",
    "スタッフコード",
    "氏名",
    "出勤日数",
    "所定労働時間",
    "実労働時間",
    "時間外労働時間",
    "深夜労働時間",
]


def create_sample_excel(output_path: str = "sample_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "勤怠サマリー"

    # ヘッダースタイル
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(HEADER_ROW, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for row_idx, (store, name, overtime) in enumerate(SAMPLE_DATA, start=2):
        ws.cell(row=row_idx, column=1, value=store)
        ws.cell(row=row_idx, column=2, value=f"S{row_idx - 1:03d}")
        ws.cell(row=row_idx, column=3, value=name)
        ws.cell(row=row_idx, column=4, value=20)
        ws.cell(row=row_idx, column=5, value="160:00")
        ws.cell(row=row_idx, column=6, value="172:30")
        ws.cell(row=row_idx, column=7, value=overtime)
        ws.cell(row=row_idx, column=8, value="0:00")

    # 列幅調整
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 16
    for col in ["B", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 14

    wb.save(output_path)
    print(f"サンプルデータを作成しました: {output_path}")
    print(f"  店舗数: {len(set(d[0] for d in SAMPLE_DATA))}")
    print(f"  スタッフ数: {len(SAMPLE_DATA)}")


if __name__ == "__main__":
    create_sample_excel()
